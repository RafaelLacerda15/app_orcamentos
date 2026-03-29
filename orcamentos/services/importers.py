import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..models import Supplier


@dataclass
class ImportResult:
    imported: int
    skipped: int
    errors: list[str]


LEGACY_CANDIDATES: dict[str, tuple[str, ...]] = {
    "name": ("nome", "name", "fornecedor", "contato"),
    "company": ("empresa", "company", "razao_social", "razaosocial"),
    "phone": ("telefone", "phone", "whatsapp", "celular"),
    "email": ("email", "e_mail", "mail"),
    "notes": ("observacoes", "obs", "notes"),
}

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "nome",
        "nome completo",
        "name",
        "fornecedor",
        "contato",
        "responsavel",
        "representante",
        "pessoa contato",
    ),
    "company": (
        "empresa",
        "company",
        "razao social",
        "razao",
        "nome fantasia",
        "fantasia",
        "organizacao",
    ),
    "phone": (
        "telefone",
        "phone",
        "celular",
        "whatsapp",
        "whats",
        "fone",
        "tel",
    ),
    "email": (
        "email",
        "e mail",
        "mail",
        "correio eletronico",
        "email comercial",
    ),
    "notes": (
        "observacoes",
        "observacao",
        "obs",
        "notes",
        "anotacoes",
        "comentarios",
        "detalhes",
    ),
}

EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def normalize_phone(value: str | None) -> str | None:
    if not value:
        return None
    digits = re.sub(r"\D", "", value)
    if not digits:
        return None
    if len(digits) == 10 or len(digits) == 11:
        return f"+55{digits}"
    if digits.startswith("55"):
        return f"+{digits}"
    return f"+{digits}"


def _normalize_text(value: str) -> str:
    plain = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    plain = plain.lower().strip().replace("-", "_").replace(" ", "_")
    return plain


def _normalize_header(value: str) -> str:
    plain = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    plain = re.sub(r"[^a-zA-Z0-9]+", " ", plain).lower().strip()
    return re.sub(r"\s+", " ", plain)


def _pick_value(row: dict[str, Any], candidates: tuple[str, ...]) -> str:
    for candidate in candidates:
        if candidate in row and row[candidate] is not None:
            text = str(row[candidate]).strip()
            if text:
                return text
    return ""


def _looks_like_email(value: str) -> bool:
    return bool(EMAIL_PATTERN.match(value.strip().lower()))


def _looks_like_phone(value: str) -> bool:
    if "@" in value:
        return False
    digits = re.sub(r"\D", "", value)
    return 8 <= len(digits) <= 14


def _score_header_for_field(header: str, field: str) -> int:
    if not header:
        return 0

    header_compact = header.replace(" ", "")
    header_tokens = set(header.split())
    best = 0

    for alias in HEADER_ALIASES[field]:
        alias_norm = _normalize_header(alias)
        if not alias_norm:
            continue

        alias_compact = alias_norm.replace(" ", "")
        alias_tokens = alias_norm.split()

        if header == alias_norm:
            best = max(best, 130)
            continue
        if header_compact == alias_compact:
            best = max(best, 120)
            continue
        if header.startswith(alias_norm) or header.endswith(alias_norm):
            best = max(best, 106)
            continue
        if alias_norm in header:
            best = max(best, 94)
            continue
        if all(token in header_tokens for token in alias_tokens):
            best = max(best, 86)
            continue

        ratio = SequenceMatcher(None, header, alias_norm).ratio()
        if ratio >= 0.82:
            best = max(best, int(ratio * 90))
        elif ratio >= 0.72:
            best = max(best, int(ratio * 72))

    if field == "email" and ("email" in header_tokens or "mail" in header_tokens):
        best = max(best, 105)
    if field == "phone" and header_tokens.intersection({"tel", "fone", "whats", "whatsapp", "celular", "telefone"}):
        best = max(best, 102)
    if field == "name" and header_tokens.intersection({"nome", "contato", "responsavel"}):
        best = max(best, 98)
    if field == "company" and header_tokens.intersection({"empresa", "razao", "fantasia", "company"}):
        best = max(best, 98)
    if field == "notes" and header_tokens.intersection({"obs", "observacao", "observacoes", "nota", "detalhes"}):
        best = max(best, 98)

    return best


def _sample_column_values(rows: list[dict[str, Any]], header: str, limit: int = 30) -> list[str]:
    values: list[str] = []
    for row in rows[:limit]:
        if header not in row or row[header] is None:
            continue
        text = str(row[header]).strip()
        if text:
            values.append(text)
    return values


def _infer_by_value_pattern(
    rows: list[dict[str, Any]],
    headers: list[str],
    used_headers: set[str],
    predicate,
    min_ratio: float,
    min_hits: int = 2,
) -> str | None:
    best_header = None
    best_ratio = 0.0

    for header in headers:
        if header in used_headers:
            continue
        sample = _sample_column_values(rows, header)
        if not sample:
            continue
        hits = sum(1 for value in sample if predicate(value))
        ratio = hits / len(sample)
        if hits >= min_hits and ratio >= min_ratio and ratio > best_ratio:
            best_header = header
            best_ratio = ratio

    return best_header


def _infer_name_by_values(rows: list[dict[str, Any]], headers: list[str], used_headers: set[str]) -> str | None:
    best_header = None
    best_score = 0.0

    for header in headers:
        if header in used_headers:
            continue

        sample = _sample_column_values(rows, header)
        if not sample:
            continue

        good = 0
        for value in sample:
            has_letters = any(char.isalpha() for char in value)
            has_long_digits = len(re.sub(r"\D", "", value)) >= 8
            if has_letters and not has_long_digits and not _looks_like_email(value):
                good += 1
        score = good / len(sample)
        if score >= 0.65 and score > best_score:
            best_header = header
            best_score = score

    return best_header


def infer_column_mapping(rows: list[dict[str, Any]]) -> dict[str, str]:
    if not rows:
        return {}

    headers = [str(key).strip() for key in rows[0].keys() if key is not None and str(key).strip()]
    normalized_by_header = {header: _normalize_header(header) for header in headers}

    scored_candidates: list[tuple[int, str, str]] = []
    for field in HEADER_ALIASES:
        for header, normalized in normalized_by_header.items():
            score = _score_header_for_field(normalized, field)
            if score >= 62:
                scored_candidates.append((score, field, header))

    mapping: dict[str, str] = {}
    used_headers: set[str] = set()
    used_fields: set[str] = set()

    for score, field, header in sorted(scored_candidates, key=lambda item: item[0], reverse=True):
        if field in used_fields or header in used_headers:
            continue
        mapping[field] = header
        used_fields.add(field)
        used_headers.add(header)

    if "email" not in mapping:
        email_header = _infer_by_value_pattern(rows, headers, used_headers, _looks_like_email, min_ratio=0.7)
        if email_header:
            mapping["email"] = email_header
            used_headers.add(email_header)

    if "phone" not in mapping:
        phone_header = _infer_by_value_pattern(rows, headers, used_headers, _looks_like_phone, min_ratio=0.65)
        if phone_header:
            mapping["phone"] = phone_header
            used_headers.add(phone_header)

    if "name" not in mapping:
        name_header = _infer_name_by_values(rows, headers, used_headers)
        if name_header:
            mapping["name"] = name_header

    return mapping


def _canonical_row(row: dict[str, Any], column_mapping: dict[str, str]) -> dict[str, str]:
    normalized_row = {_normalize_text(str(key)): value for key, value in row.items()}
    canonical: dict[str, str] = {}

    for field, candidates in LEGACY_CANDIDATES.items():
        value = ""
        mapped_header = column_mapping.get(field)
        if mapped_header and mapped_header in row and row[mapped_header] is not None:
            text = str(row[mapped_header]).strip()
            if text:
                value = text

        if not value:
            value = _pick_value(normalized_row, candidates)

        canonical[field] = value

    return canonical


def _read_csv(file_bytes: bytes) -> list[dict[str, Any]]:
    text = None
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Nao foi possivel ler o arquivo CSV.")

    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _read_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(col).strip() if col is not None else "" for col in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        row_dict = {}
        for index, value in enumerate(row):
            if index < len(headers):
                row_dict[headers[index]] = value
        data_rows.append(row_dict)
    return data_rows


def _make_dedupe_keys(phone: str | None, email: str | None) -> tuple[str | None, str | None]:
    phone_key = f"phone:{phone}" if phone else None
    email_key = f"email:{email.lower()}" if email else None
    return phone_key, email_key


def parse_rows_from_file(filename: str, file_bytes: bytes) -> list[dict[str, Any]]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return _read_csv(file_bytes)
    if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_xlsx(file_bytes)
    raise ValueError("Formato nao suportado. Use CSV ou XLSX.")


def import_suppliers_from_rows(rows: list[dict[str, Any]], existing_suppliers: list[Supplier]) -> tuple[list[Supplier], ImportResult]:
    column_mapping = infer_column_mapping(rows)

    existing_keys = set()
    for supplier in existing_suppliers:
        phone_key, email_key = _make_dedupe_keys(supplier.phone, supplier.email)
        if phone_key:
            existing_keys.add(phone_key)
        if email_key:
            existing_keys.add(email_key)

    created: list[Supplier] = []
    in_batch_keys = set()
    skipped = 0
    errors: list[str] = []

    for line_number, raw_row in enumerate(rows, start=2):
        row = _canonical_row(raw_row, column_mapping)
        name = row["name"]
        phone = normalize_phone(row["phone"])
        email = row["email"].lower() if row["email"] else None

        if not name:
            skipped += 1
            errors.append(f"Linha {line_number}: nome vazio.")
            continue

        if not phone and not email:
            skipped += 1
            errors.append(f"Linha {line_number}: informe telefone ou email.")
            continue

        phone_key, email_key = _make_dedupe_keys(phone, email)
        duplicate = False
        for key in (phone_key, email_key):
            if key and (key in existing_keys or key in in_batch_keys):
                duplicate = True
                break
        if duplicate:
            skipped += 1
            continue

        created_supplier = Supplier(
            name=name,
            company=row["company"] or None,
            phone=phone,
            email=email,
            notes=row["notes"] or None,
        )
        created.append(created_supplier)

        if phone_key:
            in_batch_keys.add(phone_key)
        if email_key:
            in_batch_keys.add(email_key)

    result = ImportResult(imported=len(created), skipped=skipped, errors=errors)
    return created, result
